#!/usr/bin/env python3
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series, BarChart, ScatterChart
from openpyxl.styles import Alignment

# Local imports
import sys
_PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))
from scripts.lib.config import load_config
from scripts.lib.paths import PathResolver, find_reference_cif, find_vina_pose
from scripts.lib.pose_pipeline import run_pose_benchmark


@dataclass
class VinaMetrics:
    top1: float | None
    best10: float | None
    avg10: float | None


def compute_vina_metrics(pdbid: str, resolver: PathResolver, ref: Path, vina_pose: Path, *, k: int = 10) -> VinaMetrics:
    # Top-1 (first model only)
    top1_res = run_pose_benchmark(
        pdbid=pdbid,
        resolver=resolver,
        project_root=resolver.config.base_dir,
        ref_path=ref,
        pred_path=vina_pose,
        pose_count=1,
        include_h=False,
        include_small=False,
        enable_pocket=True,
        pocket_radius=5.0,
        capture_full=False,
    )
    top1 = float(top1_res.get("summary", {}).get("best_pose", {}).get("rmsd_locked_global", np.nan))

    # Top-K (evaluate first K models in the file)
    k = max(1, k)
    multi_res = run_pose_benchmark(
        pdbid=pdbid,
        resolver=resolver,
        project_root=resolver.config.base_dir,
        ref_path=ref,
        pred_path=vina_pose,
        pose_count=k,
        include_h=False,
        include_small=False,
        enable_pocket=True,
        pocket_radius=5.0,
        capture_full=True,
    )

    # Extract per-pose best RMSD from details
    pose_to_vals: Dict[int, List[float]] = {}
    for row in multi_res.get("details", []):
        if row.get("record_type") != "ligand_match":
            continue
        pi = int(row.get("pose_index", 1))
        val = row.get("rmsd_locked_global")
        try:
            v = float(val)
            if np.isfinite(v):
                pose_to_vals.setdefault(pi, []).append(v)
        except Exception:
            continue

    per_pose_best: List[float] = []
    for pi in sorted(pose_to_vals):
        if pose_to_vals[pi]:
            per_pose_best.append(min(pose_to_vals[pi]))

    best10 = float(np.nanmin(per_pose_best)) if per_pose_best else np.nan
    avg10 = float(np.nanmean(per_pose_best)) if per_pose_best else np.nan
    return VinaMetrics(top1=top1, best10=best10, avg10=avg10)


def success_rate(vals: pd.Series, thr: float) -> float:
    s = pd.to_numeric(vals, errors="coerce")
    s = s.replace([np.inf, -np.inf], np.nan).dropna()
    if len(s) == 0:
        return 0.0
    return float((s <= thr).mean())


def build_cdf(values: pd.Series, grid: List[float]) -> List[float]:
    s = pd.to_numeric(values, errors="coerce")
    s = s.replace([np.inf, -np.inf], np.nan).dropna().values
    out = []
    for t in grid:
        out.append(float(np.mean(s <= t)) if len(s) else 0.0)
    return out


def add_line_chart(ws, data_range, categories_range, title: str, series_names: List[str]):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.y_axis.title = "Fraction"
    chart.x_axis.title = "RMSD threshold (Å)"
    data = Reference(ws, min_col=data_range[0], min_row=data_range[1], max_col=data_range[2], max_row=data_range[3])
    cats = Reference(ws, min_col=categories_range[0], min_row=categories_range[1], max_col=categories_range[2], max_row=categories_range[3])
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"H2")


def add_bar_chart(ws, data_range, categories_range, title: str):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = "Success rate"
    chart.x_axis.title = "Threshold (Å)"
    data = Reference(ws, min_col=data_range[0], min_row=data_range[1], max_col=data_range[2], max_row=data_range[3])
    cats = Reference(ws, min_col=categories_range[0], min_row=categories_range[1], max_col=categories_range[2], max_row=categories_range[3])
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"H2")


def add_scatter(ws, x_col, y_col, min_row, max_row, title: str, x_title: str, y_title: str):
    chart = ScatterChart()
    chart.title = title
    chart.style = 2
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    xvalues = Reference(ws, min_col=x_col, min_row=min_row, max_row=max_row)
    yvalues = Reference(ws, min_col=y_col, min_row=min_row, max_row=max_row)
    series = Series(yvalues, xvalues, title="Pairs")
    chart.series.append(series)
    ws.add_chart(chart, "H2")


def main():
    ap = argparse.ArgumentParser(description="Generate Boltz vs Vina report with plots")
    ap.add_argument("--analysis-csv", default=None, help="Path to benchmark_<timestamp>.csv (default: latest under analysis/")
    ap.add_argument("--k", type=int, default=10, help="Top-K for Vina metrics (default: 10)")
    args = ap.parse_args()

    config = load_config()
    analysis_dir = config.paths.analysis_root
    if args.analysis_csv:
        csv_path = Path(args.analysis_csv)
    else:
        candidates = sorted(analysis_dir.glob("benchmark_*.csv"), key=lambda p: p.stat().st_mtime)
        if not candidates:
            raise SystemError("No benchmark_*.csv found in analysis/")
        csv_path = candidates[-1]

    df = pd.read_csv(csv_path)
    # Collect paired proteins
    pdb_ids = sorted(df["pdb_id"].unique())

    # Build per-protein Boltz values
    boltz = df[df["method"] == "boltz"].set_index("pdb_id")["rmsd_global"]

    # Vina metrics per protein
    vina_resolver = PathResolver(config, preds=config.paths.vina_preds, analysis_dir=analysis_dir)
    rows: List[Dict[str, object]] = []
    for pid in pdb_ids:
        ref = find_reference_cif(pid, vina_resolver.refs_root)
        if ref is None:
            continue
        vina_pose = find_vina_pose(pid, vina_resolver.preds_root)
        if vina_pose is None:
            continue
        metrics = compute_vina_metrics(pid, vina_resolver, ref, vina_pose, k=args.k)
        rows.append(
            {
                "pdb_id": pid,
                "boltz": float(boltz.get(pid, np.nan)) if pid in boltz.index else np.nan,
                "vina_top1": metrics.top1,
                "vina_best10": metrics.best10,
                "vina_avg10": metrics.avg10,
            }
        )

    out_df = pd.DataFrame(rows).dropna(subset=["boltz", "vina_top1", "vina_best10", "vina_avg10"], how="all")
    out_dir = analysis_dir / "report"
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "boltz_vs_vina_report.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        # Data sheet
        out_df.to_excel(writer, sheet_name="Data", index=False)
        wb = writer.book
        ws_data = writer.sheets["Data"]
        ws_data.freeze_panes = "A2"

        # Summary sheet
        ws_sum = wb.create_sheet("Summary")
        ws_sum.append(["metric", "boltz", "vina_top1", "vina_best10", "vina_avg10"])

        def add_row(name, func):
            vals = [func(out_df["boltz"]), func(out_df["vina_top1"]), func(out_df["vina_best10"]), func(out_df["vina_avg10"])]
            ws_sum.append([name, *[float(v) if v == v else "" for v in vals]])

        add_row("mean_rmsd", lambda s: pd.to_numeric(s, errors="coerce").mean())
        add_row("median_rmsd", lambda s: pd.to_numeric(s, errors="coerce").median())
        add_row("std_rmsd", lambda s: pd.to_numeric(s, errors="coerce").std())
        for thr in [1.0, 2.0, 3.0]:
            add_row(f"success_rate_<=_{thr}A", lambda s, t=thr: success_rate(pd.to_numeric(s, errors="coerce"), t))

        # Thresholds sheet + bar chart
        ws_thr = wb.create_sheet("Thresholds")
        ws_thr.append(["threshold", "boltz", "vina_top1", "vina_best10", "vina_avg10"])
        thresholds = [0.5, 1.0, 2.0, 3.0, 4.0]
        for t in thresholds:
            ws_thr.append([t, success_rate(out_df["boltz"], t), success_rate(out_df["vina_top1"], t), success_rate(out_df["vina_best10"], t), success_rate(out_df["vina_avg10"], t)])
        add_bar_chart(ws_thr, data_range=(2, 1, 5, 1 + len(thresholds)), categories_range=(1, 2, 1, 1 + len(thresholds)), title="Success rates by threshold")

        # CDF sheet + line chart
        ws_cdf = wb.create_sheet("CDF")
        ws_cdf.append(["threshold", "boltz", "vina_top1", "vina_best10", "vina_avg10"])
        grid = [round(x, 2) for x in np.linspace(0.0, 10.0, 41)]
        b_cdf = build_cdf(out_df["boltz"], grid)
        t1_cdf = build_cdf(out_df["vina_top1"], grid)
        b10_cdf = build_cdf(out_df["vina_best10"], grid)
        a10_cdf = build_cdf(out_df["vina_avg10"], grid)
        for i, t in enumerate(grid):
            ws_cdf.append([t, b_cdf[i], t1_cdf[i], b10_cdf[i], a10_cdf[i]])
        add_line_chart(ws_cdf, data_range=(2, 1, 5, 1 + len(grid)), categories_range=(1, 2, 1, 1 + len(grid)), title="CDF of locked global RMSD", series_names=["boltz", "vina_top1", "vina_best10", "vina_avg10"])

        # Paired sheet + scatter (Boltz vs Vina best-10)
        ws_pair = wb.create_sheet("Paired")
        ws_pair.append(["boltz", "vina_best10"])
        for _, r in out_df.dropna(subset=["boltz", "vina_best10"]).iterrows():
            ws_pair.append([float(r["boltz"]), float(r["vina_best10"])])
        add_scatter(ws_pair, x_col=1, y_col=2, min_row=2, max_row=ws_pair.max_row, title="Boltz vs Vina (best-10)", x_title="Boltz RMSD", y_title="Vina best-10 RMSD")

        # Autofit widths
        for ws in [ws_data, ws_sum, ws_thr, ws_cdf, ws_pair]:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value)
                    except Exception:
                        val = ""
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(40, max(10, max_len + 2))

    # Also write a brief Markdown narrative
    md_path = out_dir / "boltz_vs_vina_report.md"
    b_med = float(pd.to_numeric(out_df["boltz"]).median())
    t1_med = float(pd.to_numeric(out_df["vina_top1"]).median())
    b10_med = float(pd.to_numeric(out_df["vina_best10"]).median())
    a10_med = float(pd.to_numeric(out_df["vina_avg10"]).median())

    lines = []
    lines.append("# Boltz vs Vina Report\n")
    lines.append(f"Source CSV: `{csv_path}`\n")
    lines.append("## Summary\n")
    lines.append(f"- Proteins analyzed: {len(out_df)}\n")
    lines.append(f"- Median locked global RMSD (Å): Boltz {b_med:.3f}, Vina top-1 {t1_med:.3f}, Vina best-10 {b10_med:.3f}, Vina avg-10 {a10_med:.3f}\n")
    lines.append("- See `boltz_vs_vina_report.xlsx` for charts (CDF, thresholds, paired scatter) and complete tables.\n")
    md_path.write_text("\n".join(lines))

    print("Wrote", xlsx_path)
    print("Wrote", md_path)


if __name__ == "__main__":
    main()
